import csv
import io
import logging
import os
from datetime import date, datetime, timezone
from decimal import Decimal, InvalidOperation

from fastapi import HTTPException, UploadFile, status
from sqlalchemy.orm import Session

from openpyxl import load_workbook

from app.core.database import settings
from app.models.timesheet import Timesheet
from app.models.timesheet_entry import TimesheetEntry
from app.services.employee_client import get_employee_details


logger = logging.getLogger(__name__)


ALLOWED_EXTENSIONS = {".xlsx", ".csv"}

REQUIRED_COLUMNS = {
    "employee_id",
    "employee_name",
    "regular_hours",
    "overtime_hours",
}


def _validate_file(file: UploadFile) -> str:
    """
    Validate uploaded timesheet file.
    """

    if not file.filename:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Timesheet file is required.",
        )

    extension = os.path.splitext(file.filename)[1].lower()

    if extension not in ALLOWED_EXTENSIONS:
        logger.warning(
            "Invalid timesheet file type | filename=%s",
            file.filename,
        )

        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Only .xlsx and .csv timesheet files are supported.",
        )

    return extension


def _parse_decimal(
    value,
    field_name: str,
    row_number: int,
) -> Decimal:
    """
    Safely convert a spreadsheet value into Decimal.
    """

    if value is None or str(value).strip() == "":
        return Decimal("0.00")

    try:
        decimal_value = Decimal(str(value))

        if decimal_value < 0:
            raise ValueError

        return decimal_value.quantize(
            Decimal("0.01")
        )

    except (InvalidOperation, ValueError):

        logger.warning(
            "Invalid numeric value | row=%s | field=%s | value=%s",
            row_number,
            field_name,
            value,
        )

        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=(
                f"Invalid {field_name} value "
                f"at row {row_number}."
            ),
        )


def _validate_columns(columns: list[str]):
    """
    Validate required spreadsheet columns.
    """

    normalized_columns = {
        column.strip().lower()
        for column in columns
        if column
    }

    missing_columns = REQUIRED_COLUMNS - normalized_columns

    if missing_columns:

        logger.warning(
            "Missing timesheet columns | missing=%s",
            missing_columns,
        )

        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=(
                "Missing required columns: "
                + ", ".join(sorted(missing_columns))
            ),
        )


def _parse_csv(
    file_content: bytes,
) -> list[dict]:
    """
    Parse CSV timesheet content.
    """

    try:
        text = file_content.decode("utf-8-sig")

        reader = csv.DictReader(
            io.StringIO(text)
        )

        if not reader.fieldnames:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="CSV file is empty.",
            )

        _validate_columns(
            reader.fieldnames
        )

        rows = []

        for row in reader:
            normalized_row = {
                key.strip().lower(): value
                for key, value in row.items()
                if key
            }

            rows.append(normalized_row)

        return rows

    except UnicodeDecodeError:

        logger.exception(
            "Unable to decode CSV timesheet"
        )

        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="CSV file must use UTF-8 encoding.",
        )


def _parse_excel(
    file_content: bytes,
) -> list[dict]:
    """
    Parse XLSX timesheet content.
    """

    try:

        workbook = load_workbook(
            filename=io.BytesIO(file_content),
            read_only=True,
            data_only=True,
        )

        worksheet = workbook.active

        rows = worksheet.iter_rows(
            values_only=True
        )

        try:
            header_row = next(rows)
        except StopIteration:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Excel file is empty.",
            )

        headers = [
            str(header).strip().lower()
            if header is not None
            else ""
            for header in header_row
        ]

        _validate_columns(headers)

        result = []

        for row in rows:

            if not any(
                value is not None
                for value in row
            ):
                continue

            row_data = {}

            for index, header in enumerate(headers):

                if not header:
                    continue

                value = (
                    row[index]
                    if index < len(row)
                    else None
                )

                row_data[header] = value

            result.append(row_data)

        workbook.close()

        return result

    except HTTPException:
        raise

    except Exception:

        logger.exception(
            "Failed to parse Excel timesheet"
        )

        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Unable to read Excel timesheet file.",
        )


def _parse_file(
    file_content: bytes,
    extension: str,
) -> list[dict]:

    if extension == ".csv":
        return _parse_csv(file_content)

    if extension == ".xlsx":
        return _parse_excel(file_content)

    raise HTTPException(
        status_code=status.HTTP_400_BAD_REQUEST,
        detail="Unsupported timesheet file.",
    )


def _validate_period(
    period_start: date,
    period_end: date,
):
    if period_start > period_end:

        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Period start date cannot be after period end date.",
        )


def _validate_rows(
    rows: list[dict],
) -> list[dict]:
    """
    Validate and normalize parsed timesheet rows.
    """

    if not rows:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Timesheet contains no employee records.",
        )

    validated_rows = []

    employee_ids = set()

    for row_number, row in enumerate(
        rows,
        start=2,
    ):

        employee_id_value = row.get(
            "employee_id"
        )

        employee_name = str(
            row.get("employee_name") or ""
        ).strip()

        if not employee_id_value:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=(
                    f"Employee ID is required "
                    f"at row {row_number}."
                ),
            )

        try:
            employee_id = int(
                employee_id_value
            )
        except (TypeError, ValueError):

            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=(
                    f"Invalid employee ID "
                    f"at row {row_number}."
                ),
            )

        if employee_id in employee_ids:

            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=(
                    f"Duplicate employee ID "
                    f"{employee_id} at row {row_number}."
                ),
            )

        employee_ids.add(employee_id)

        if not employee_name:

            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=(
                    f"Employee name is required "
                    f"at row {row_number}."
                ),
            )

        regular_hours = _parse_decimal(
            row.get("regular_hours"),
            "regular_hours",
            row_number,
        )

        overtime_hours = _parse_decimal(
            row.get("overtime_hours"),
            "overtime_hours",
            row_number,
        )

        total_hours = (
            regular_hours +
            overtime_hours
        ).quantize(
            Decimal("0.01")
        )

        validated_rows.append(
            {
                "employee_id": employee_id,
                "employee_name": employee_name,
                "regular_hours": regular_hours,
                "overtime_hours": overtime_hours,
                "total_hours": total_hours,
            }
        )

    return validated_rows




def upload_timesheet(
    db: Session,
    client_id: int,
    period_start: date,
    period_end: date,
    file: UploadFile,
    uploaded_by: int | None = None,
):
    """
    Upload and process a client timesheet.
    """

    logger.info(
        "Processing timesheet upload | client_id=%s | "
        "period=%s to %s | filename=%s",
        client_id,
        period_start,
        period_end,
        file.filename,
    )

    _validate_period(
        period_start,
        period_end,
    )

    extension = _validate_file(file)

    file_content = file.file.read()

    if not file_content:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Uploaded timesheet file is empty.",
        )

    rows = _parse_file(
        file_content,
        extension,
    )

    validated_rows = _validate_rows(
        rows
    )

    os.makedirs(
        settings.UPLOAD_DIR,
        exist_ok=True,
    )

    safe_filename = (
        f"{datetime.now(timezone.utc).strftime('%Y%m%d%H%M%S')}_"
        f"{os.path.basename(file.filename)}"
    )

    file_path = os.path.join(
        settings.UPLOAD_DIR,
        safe_filename,
    )

    try:

        with open(
            file_path,
            "wb",
        ) as destination:

            destination.write(
                file_content
            )

        timesheet = Timesheet(
            client_id=client_id,
            period_start=period_start,
            period_end=period_end,
            file_name=file.filename,
            file_path=file_path,
            status="UPLOADED",
            uploaded_by=uploaded_by,
        )

        db.add(timesheet)
        db.flush()

        for row in validated_rows:

            employee = get_employee_details(
                row["employee_id"]
            )

            if not employee.get("is_active", False):

                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail=(
                        f"Employee "
                        f"{row['employee_id']} is inactive."
                    ),
                )

            if employee.get("client_id") != client_id:

                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail=(
                        f"Employee "
                        f"{row['employee_id']} does not belong "
                        f"to client {client_id}."
                    ),
                )

            hourly_rate = Decimal(
                str(employee["hourly_rate"])
            ).quantize(
                Decimal("0.01")
            )

            amount = (
                row["total_hours"] *
                hourly_rate
            ).quantize(
                Decimal("0.01")
            )

            entry = TimesheetEntry(
                timesheet_id=timesheet.id,
                employee_id=row["employee_id"],
                employee_name=(
                    f"{employee['first_name']} "
                    f"{employee['last_name']}"
                ).strip(),
                regular_hours=row["regular_hours"],
                overtime_hours=row["overtime_hours"],
                total_hours=row["total_hours"],
                hourly_rate=hourly_rate,
                amount=amount,
            )

            db.add(entry)

        db.commit()
        db.refresh(timesheet)

        logger.info(
            "Timesheet uploaded successfully | "
            "timesheet_id=%s | client_id=%s | entries=%s",
            timesheet.id,
            client_id,
            len(validated_rows),
        )

        return timesheet

    except HTTPException:
        db.rollback()

        if os.path.exists(file_path):
            os.remove(file_path)

        raise

    except Exception:

        db.rollback()

        if os.path.exists(file_path):
            os.remove(file_path)

        logger.exception(
            "Failed to upload timesheet | client_id=%s",
            client_id,
        )

        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to process timesheet.",
        )


def get_timesheets(
    db: Session,
):
    """
    Return all timesheets.
    """

    logger.info(
        "Fetching all timesheets"
    )

    return (
        db.query(Timesheet)
        .order_by(
            Timesheet.created_at.desc()
        )
        .all()
    )


def get_timesheet(
    db: Session,
    timesheet_id: int,
):
    """
    Return a single timesheet.
    """

    timesheet = (
        db.query(Timesheet)
        .filter(
            Timesheet.id == timesheet_id
        )
        .first()
    )

    if not timesheet:

        logger.warning(
            "Timesheet not found | id=%s",
            timesheet_id,
        )

        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Timesheet not found.",
        )

    return timesheet


def approve_timesheet(
    db: Session,
    timesheet_id: int,
    approved_by: int | None = None,
):
    """
    Approve an uploaded timesheet.
    """

    timesheet = get_timesheet(
        db,
        timesheet_id,
    )

    if timesheet.status not in {
        "UPLOADED",
        "UNDER_REVIEW",
    }:

        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=(
                f"Timesheet cannot be approved "
                f"from status {timesheet.status}."
            ),
        )

    timesheet.status = "APPROVED"
    timesheet.approved_by = approved_by
    timesheet.approved_at = datetime.now(
        timezone.utc
    )

    db.commit()
    db.refresh(timesheet)

    logger.info(
        "Timesheet approved | id=%s | approved_by=%s",
        timesheet_id,
        approved_by,
    )

    return {
        "id": timesheet.id,
        "status": timesheet.status,
        "message": "Timesheet approved successfully.",
    }


def reject_timesheet(
    db: Session,
    timesheet_id: int,
):
    """
    Reject a timesheet.
    """

    timesheet = get_timesheet(
        db,
        timesheet_id,
    )

    if timesheet.status not in {
        "UPLOADED",
        "UNDER_REVIEW",
    }:

        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=(
                f"Timesheet cannot be rejected "
                f"from status {timesheet.status}."
            ),
        )

    timesheet.status = "REJECTED"

    db.commit()
    db.refresh(timesheet)

    logger.info(
        "Timesheet rejected | id=%s",
        timesheet_id,
    )

    return {
        "id": timesheet.id,
        "status": timesheet.status,
        "message": "Timesheet rejected successfully.",
    }